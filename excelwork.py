import openpyxl

'''
Simple definitions:
Workbook - .xlsx extension
Sheets / Worksheets - individual "pages" in a workbook
Columns (letters) , Rows (numbers), Cells (Column+Row)
'''


def excel_notes():
    workbook = openpyxl.load_workbook('Data\\example.xlsx')
    # sheet1 = workbook.get_sheet_by_name('Sheet1') # deprecated
    sheet1 = workbook['Sheet1']
    print(workbook.get_sheet_names)
    cell = sheet1['A1']
    print(cell.value)
    print(str(cell.value))
    print(sheet1.cell(row=1, column=2).value)
    for i in range(1,8):
        print(i, sheet1.cell(row=i, column=2).value)


def more_excel_notes():
    wb = openpyxl.Workbook()
    print(wb.get_sheet_names())
    sheet = wb['Sheet']
    sheet['A1'] = 42
    sheet['A2'] = 'Hello'
    sheet2 = wb.create_sheet()
    sheet2.title = 'NewSheet'
    sheet0 = wb.create_sheet(index=0, title='Sheet0')
    # good practice - when opening a worksheet and editing it, save it to a new file
    wb.save('Data\\new_workbook.xlsx')


def main():
    excel_notes()


if __name__ == '__main__':
    main()
